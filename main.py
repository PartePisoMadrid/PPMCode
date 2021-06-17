#!/home/ebotiab/anaconda3/bin/python

# -*- coding: utf-8 -*-
"""
Created on Fri Sep 11 09:59:29 2020

@author: PartePisoMadrid
"""
from googleapiclient.discovery import build

from utils.google_apis import GLogin
from utils.google_apis import GDrive
from utils.google_apis import GSheets
from utils.google_apis import GContacts
from utils.FileConvert import gsheet2csv
from utils.FileConvert import csv2excel
from utils import TKinterTools

from openpyxl import load_workbook
import smtplib

import pandas as pd
import datetime
import subprocess, sys
import os
import re

## CONSTANTS
# if modifying these scopes, delete the file token.json
SCOPES = ['https://www.googleapis.com/auth/drive https://www.googleapis.com/auth/contacts']
# change these paths as appropriate
CREDENTIALS_PATH = '/home/ebotiab/Desktop/parte/credentials/'
CONTACTS_PATH = '/home/ebotiab/Desktop/parte/mycontacts.txt'
URL_FILE_PATH = "Parte_files/url_parte.txt"
SHMESS_FILE_PATH = "Parte_files/sharing_parte_message.txt"
# template Parte ID
FILE_ID = "1KKsoosB88DlOIHgBkd4K9iQxbPeUvWmvB_u9RhuJ2N8"
# ID of the folder where new Parte will be located if needed
FOLDER_ID = "1xYJf9E9oRs3cvtdSNFLo3qolo8438lR4"
# name of the group to which Partes will be shared
#PISO_GROUP_NAME = "PisoMadrid2020-21"
PISO_GROUP_NAME = "Verano2020-21"
# mapping of month number to spanish month name
SPANISH_MONTH = {1:"enero", 2:"febrero", 3:"marzo", 4:"abril", 5:"mayo", 6:"junio", 7:"julio", 8:"agosto", 9:"septiembre", 10:"octubre",11:"noviembre", 12:"diciembre"}

def printParte():
    """
    Downloads and prints today's Parte
    """
    fileToRead = open(urlFilePath, "r")
    parteUrl = fileToRead.read()
    fileToRead.close()

    # download Parte google sheet to a csv
    parteCsv = gsheet2csv(CREDENTIALS_PATH, parteUrl, 5, 'Parte_files/resumen.csv') 
    # convert csv downloaded to excel
    parteDf = csv2excel("Parte_files/resumen.csv", 'Parte_files/resumen_data.xlsx', 'Sheet1') 
    # store date
    weekDay = parteDf.columns[2]

    # load workbook 
    wb_to_write = load_workbook('Parte_files/RESUMEN_PARTE.XLSX')
    # get sheets of the woorkbooks
    sheet_to_read = load_workbook('Parte_files/resumen_data.xlsx')['Sheet1']
    sheet_to_write = wb_to_write['Resumen_cocina']

    # update content of RESUMEN_PARTE.xlsx
    for row_read in range(2,23):
        row_write = row_read
        
        if row_write>12:
            row_write+=2
        if row_write>21:
            row_write+=1
        
        for col in range(3,8):
            if type(sheet_to_write.cell(row_write,col)).__name__ == 'MergedCell':
                continue
            sheet_to_write.cell(row_write,col).value = ""
            if type(sheet_to_read.cell(row_read,col).value)!=tuple:
                value = sheet_to_read.cell(row_read,col).value
                to_center = ["TU","TB","AF2","AF3"]
                if value in to_center:
                    value = "                       " + value
                if value=="AF2 + AF3":
                    value = "                   " + value
                sheet_to_write.cell(row_write,col).value = value
            print('row: '+str(row_read)+' col: '+str(col)+' value: '+str(sheet_to_read.cell(row_read,col).value))
    
    # save the changes in other xlsx file
    sheet_to_write.cell(1,3).value = weekDay
    wb_to_write.save("Parte_files/RESUMEN_IMPRIMIBLE.xlsx")
    
    # convert excel to pdf
    os.system("soffice --headless --convert-to pdf Parte_files/RESUMEN_IMPRIMIBLE.xlsx")
    # move to Parte_files folder
    os.rename("RESUMEN_IMPRIMIBLE.pdf", "Parte_files/RESUMEN_IMPRIMIBLE.pdf") 
    # print pdf
    os.system("lp Parte_files/RESUMEN_IMPRIMIBLE.pdf")

def createNextParte():
    """
    Creates Parte for next week
    """
    print("Creating new Parte...")
    creds = GLogin.login(SCOPES, CREDENTIALS_PATH)
    # call the Drive v3 API
    gDriveService = build('drive', 'v3', credentials=creds)
    # copy the template file
    gDriveService.files().copy(fileId=FILE_ID,
                body={"parents": [{"kind": "drive#fileLink", "id": FOLDER_ID}]}).execute()
    # get id from copy
    copyID = GDrive.title2id(gDriveService, "Copia de Parte_Coronavirus")
    # rename copy with date of next monday
    now = datetime.datetime.now()
    monday = now + datetime.timedelta(days = 7-now.weekday())
    copyTitle = "Semana "+str(monday.day)+" de "+ SPANISH_MONTH[monday.month]
    gDriveService.files().update(fileId=copyID, body={"name":copyTitle}).execute()
    # call the GSheets v4 API
    gSheetService = build('sheets', 'v4', credentials=creds)
    # modify copy by changing the dates
    dateToInsert = str(monday.day)+"/"+str(now.month)+"/"+str(now.year)
    GSheets.writeCell(gSheetService, dateToInsert, copyID, "Parte", "E3:G3", "USER_ENTERED")
    # give writer privileges to anyone with the link
    GDrive.giveAccess(gDriveService, copyID, "anyone")
    print("Parte called '"+copyTitle+"' has been successfully created")

def getContacts():
    # call the People v1 API
    creds = GLogin.login(SCOPES, CREDENTIALS_PATH)
    gContactsService = build('people', 'v1', credentials=creds)
    # get needed id and length of contacts group specified with PISO_GROUP_NAME
    pisoGroupID,pisoGroupLength = GContacts.getGroupData(gContactsService,PISO_GROUP_NAME,["resourceName","memberCount"])
    # return names and mails of the Piso contacts
    return GContacts.getContactsData(gContactsService, pisoGroupID, pisoGroupLength)

def main():
    reviseWeekDay = datetime.datetime.today().weekday()
    if reviseWeekDay==0:
        # update Parte url
        TKinterTools.writeFile(URL_FILE_PATH, 'Copy here the new Parte url')
        # create next week Parte
        createNextParte()

    # download and print Parte
    #printParte()

    if reviseWeekDay==3:
        creds = GLogin.login(SCOPES[0], CREDENTIALS_PATH)
        gDriveService = build('drive', 'v3', credentials=creds)
        #get the ID of the next week Parte
        nextMondayDate = datetime.datetime.now() + datetime.timedelta(days = 7-datetime.datetime.now().weekday())
        nextParteTitle = "Semana "+str(nextMondayDate.day)+" "+ SPANISH_MONTH[nextMondayDate.month]
        nextParteID = GDrive.title2id(gDriveService, nextParteTitle)
        # get the contacts
        mailsList = getContacts()[1]
        # share next Parte with the contacts (can be simplified by creating a Google Group)
        emailMessage = TKinterTools.writeFile(SHMESS_FILE_PATH, 'Enter the Parte sharing message', True)
        for mail in mailsList:
            GDrive.giveAccess(gDriveService, nextParteID, "user", mail, emailMessage)
        print("The Parte", nextParteTitle, "has been shared successfully with:")
        print(", ".join(mailsList))

if __name__ == '__main__':
	main()
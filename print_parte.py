#!/home/ebotiab/anaconda3/bin/python

# -*- coding: utf-8 -*-
"""
Created on Fri Sep 11 09:59:29 2020

@author: eboti
"""
import os
import pandas as pd
import datetime
import numpy as np
import subprocess, sys
import tkinter as tk

# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook
from gsheets import Sheets
#from Mail import Mail
#from PersonalData import PersonalData
# Import smtplib for the actual sending function
import smtplib
# Import the email modules we'll need
from email.message import EmailMessage

def google_sheet_to_csv(GSheetURL):
    """
    Export google sheet to csv files
    """
    sheets = Sheets.from_files('client_secrets.json', 'storage.json')
    print(GSheetURL)
    parteCsv = sheets.get(GSheetURL)
    parteCsv.sheets[5].to_csv('resumen.csv', encoding='utf-8', dialect='excel')
    ##parteCsv.sheets[1].to_csv('parte.csv', encoding='utf-8', dialect='excel')
    return parteCsv

def csvToExcel(nameCsvFile, nameExcelFile):
    """
    Read a csv file as and save it as excel file, returns the day of the week of the file
    """
    df = pd.read_csv(nameCsvFile)# Load csv
    weekDay = df.columns[2]
    writer = pd.ExcelWriter(nameExcelFile,engine='xlsxwriter')# Specify a writer
    df.to_excel(writer, 'Sheet1')# Write your DataFrame to a file
    writer.save() # Save the result
    return weekDay

def main():
    os.chdir("/home/ebotiab/Desktop/parte")#move to the Parte directory
    
    if os.path.isfile("resumen.csv"):
    	os.remove("resumen.csv")#just in case remove the resumen of the previous day

    #Update the Parte url if necessary
    reviseWeekDay = datetime.datetime.today().weekday()
    if reviseWeekDay==0:
    	ff = open("url_parte.txt", "w")
    	root= tk.Tk()
    	canvas1 = tk.Canvas(root, width = 400, height = 300)
    	canvas1.pack()
    	entry1 = tk.Entry(root)
    	canvas1.create_window(200, 140, window=entry1)
    	def getSquareRoot ():  
    		parteUrl = entry1.get()
    		ff.write(parteUrl)
    		cc = ff.close()
    		root.destroy()
    	button1 = tk.Button(text='Copy here the new Parte url', command=getSquareRoot)
    	canvas1.create_window(200, 180, window=button1)
    	root.mainloop()

    f = open("url_parte.txt", "r")
    parteUrl = f.read()
    c = f.close()
    parteCsv = google_sheet_to_csv(parteUrl)#download the Parte google sheet
    weekDay = csvToExcel("resumen.csv", 'resumen_data.xlsx') #Convert csv downloaded in excel and store date

    # Load the workbook 
    wb_to_write = load_workbook('RESUMEN_PARTE.XLSX')
    # Get the sheets of the woorkbooks
    sheet_to_read = load_workbook('resumen_data.xlsx')['Sheet1']
    sheet_to_write = wb_to_write['Resumen_cocina']

    #Update the content of RESUMEN_PARTE.xlsx
    for row_read in range(2,23):
        row_write = row_read
        
        if row_write>12:
            row_write+=2
        if row_write>21:
            row_write+=1
        
        for col in range(3,8):
            if type(sheet_to_write.cell(row_write,col)).__name__ == 'MergedCell':
                #print('CONTINUE: '+'row: '+str(row_read)+' col: '+str(col)+' value: '+str(sheet_to_read.cell(row_read,col).value))
                continue
            sheet_to_write.cell(row_write,col).value = ""
            if type(sheet_to_read.cell(row_read,col).value)!=tuple:
                value = sheet_to_read.cell(row_read,col).value
                to_center = ["TU","TB","AF2","AF3"]
                if value in to_center:
                    value = "                       "+value
                if value=="AF2 + AF3":
                    value = "                   "+value
                sheet_to_write.cell(row_write,col).value = value
            print('row: '+str(row_read)+' col: '+str(col)+' value: '+str(sheet_to_read.cell(row_read,col).value))
    
    #Save the changes in other xlsx file
    sheet_to_write.cell(1,3).value = weekDay
    wb_to_write.save("RESUMEN_IMPRIMIBLE.xlsx")
    
    #Print an excel file with the default printer
    fileLocation = "RESUMEN_IMPRIMIBLE.pdf"
    #opener ="open" if sys.platform == "darwin" else "xdg-open"
    #subprocess.call([opener, fileLocation])
    os.system("soffice --headless --convert-to pdf RESUMEN_IMPR*")
    os.system("lp "+fileLocation)

    
    #persData = PersonalData()
    #rememberMail = Mail(persData.address,persData.password,persData.port, persData.host, 'mycontacts.txt','message.txt')
    #rememberMail.send()

def futureProject():
    #os.chdir(r"/media/ebotiab/Windows/Users/eboti/Desktop/parte")
    os.chdir(r"C:\Users\eboti\Desktop\parte")
    parte = load_workbook('parte.xlsx')
    parte_sheet = parte["Parte"]
    
    names = np.zeros(30)
    count = 0
    for row in range(5,12):
        n_x = 0
        for col in range(4,25):
            if parte_sheet[row][col].value=="x":
                n_x += 1
        if n_x >18:
            names[count]==parte_sheet[row][3].value
    print(names)
                

if __name__ == '__main__':
	main()